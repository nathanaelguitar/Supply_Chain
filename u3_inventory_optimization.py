# u3_inventory_optimization.py — UNIT 3 ASSIGNMENT
# Finger Lakes Activewear (FLA) — Optimizing Production Capacity Allocation Decisions
#
# Parts covered:
#   Part 1 — Optimal production quantity WITHOUT capacity limit (critical fractile + Q*)
#   Part 2 — Optimal production quantity WITH capacity limit of 20,000 units (shadow price θ)
#   Part 3 — Risk ranking + speculative capacity (10,000 and 5,000 units)
#   Part 4 — Capacity management recommendations
#   Export — Writes all results back into the original xlsx template

import sys
from pathlib import Path
import pandas as pd
import numpy as np
from scipy.stats import norm
import shutil
import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
from u3_ingestion import load_finger_lakes

# ─────────────────────────────────────────────
# Global assignment parameters
# ─────────────────────────────────────────────
OVERAGE_PCT  = 0.10   # o = 10% of retail price
UNDERAGE_PCT = 0.30   # u = 30% of retail price

CAPACITY_PART2     = 20_000
CAPACITY_PART3A    = 10_000
CAPACITY_PART3B    =  5_000


# ─────────────────────────────────────────────
# Core newsvendor functions
# ─────────────────────────────────────────────
def compute_costs(price: float) -> tuple[float, float]:
    """Return (underage_cost, overage_cost) from retail price."""
    u = UNDERAGE_PCT * price
    o = OVERAGE_PCT  * price
    return u, o


def critical_fractile(u: float, o: float) -> float:
    return u / (u + o)


def q_star_normal(mu: float, sigma: float, u: float, o: float) -> float:
    """Q* = mu + z * sigma,  z = Phi^{-1}(CF)."""
    cf = critical_fractile(u, o)
    z  = norm.ppf(cf)
    return mu + z * sigma


# ─────────────────────────────────────────────
# Multi-item shadow-price (theta) solver
# ─────────────────────────────────────────────
def cf_with_theta(u: float, o: float, theta: float, eps: float = 1e-9) -> float:
    """CF_i(θ) = (u_i − θ) / (u_i + o_i), clamped to (eps, 1-eps)."""
    cf = (u - theta) / (u + o)
    return min(max(cf, eps), 1 - eps)


def q_with_theta_row(row: pd.Series, theta: float) -> float:
    cf = cf_with_theta(row["u"], row["o"], theta)
    z  = norm.ppf(cf)
    return row["demand_mean"] + z * row["demand_std"]


def solve_theta(df: pd.DataFrame, capacity: float,
                theta_low: float = 0.0, theta_high: float = 1_000.0,
                tol: float = 1e-6, max_iter: int = 300) -> float:
    """Bisection: find θ so that Σ Q_i(θ) == capacity."""

    def total_q(theta: float) -> float:
        return sum(q_with_theta_row(row, theta) for _, row in df.iterrows())

    lo, hi = theta_low, theta_high

    if total_q(lo) <= capacity:
        return lo

    while total_q(hi) > capacity:
        hi *= 2.0
        if hi > 1e12:
            raise RuntimeError("Could not bracket θ — capacity may be too small.")

    for _ in range(max_iter):
        mid = (lo + hi) / 2.0
        tq  = total_q(mid)
        if abs(tq - capacity) <= tol:
            return mid
        if tq > capacity:
            lo = mid
        else:
            hi = mid
    return (lo + hi) / 2.0


def allocate(df: pd.DataFrame, capacity: float) -> tuple[float, pd.DataFrame]:
    """
    Returns (theta, result_df) with columns:
        product, u, o, cf_constrained, q_constrained
    """
    theta = solve_theta(df, capacity)
    rows = []
    for _, row in df.iterrows():
        cf = cf_with_theta(row["u"], row["o"], theta)
        q  = row["demand_mean"] + norm.ppf(cf) * row["demand_std"]
        rows.append({
            "product":         row["product"],
            "u":               row["u"],
            "o":               row["o"],
            "cf_constrained":  cf,
            "q_constrained":   q,
        })
    return theta, pd.DataFrame(rows)


# ─────────────────────────────────────────────
# Risk ranking index
# ─────────────────────────────────────────────
def risk_index(u: float, o: float, mu: float, sigma: float) -> float:
    """R_i = (u_i + o_i) * (σ_i / μ_i)  — higher = riskier."""
    return (u + o) * (sigma / mu)


# ─────────────────────────────────────────────
# Pretty-printing helpers
# ─────────────────────────────────────────────
def sep(title: str = "", width: int = 72) -> None:
    if title:
        print(f"\n{'─'*width}")
        print(f"  {title}")
        print(f"{'─'*width}")
    else:
        print(f"{'─'*width}")


# ═════════════════════════════════════════════
# PART 1 — No capacity limit
# ═════════════════════════════════════════════
def part_one(df: pd.DataFrame) -> pd.DataFrame:
    sep("PART 1 — Optimal Production Quantity (No Capacity Limit)")

    results = []
    for _, row in df.iterrows():
        u, o = compute_costs(row["price"])
        cf   = critical_fractile(u, o)
        q    = q_star_normal(row["demand_mean"], row["demand_std"], u, o)
        results.append({
            "product":     row["product"],
            "price":       row["price"],
            "u":           u,
            "o":           o,
            "cf":          cf,
            "z":           norm.ppf(cf),
            "demand_mean": row["demand_mean"],
            "demand_std":  row["demand_std"],
            "q_star":      q,
        })

    res = pd.DataFrame(results)

    print(f"\n  Overage cost  = {OVERAGE_PCT*100:.0f}% of price")
    print(f"  Underage cost = {UNDERAGE_PCT*100:.0f}% of price")
    print(f"  Critical Fractile = u/(u+o) = {UNDERAGE_PCT}/{UNDERAGE_PCT+OVERAGE_PCT} = "
          f"{UNDERAGE_PCT/(UNDERAGE_PCT+OVERAGE_PCT):.4f}  (same for all products)\n")

    fmt = (
        f"  {'Product':<14} {'Price':>6} {'u':>6} {'o':>6} "
        f"{'CF':>7} {'z':>7} {'μ':>6} {'σ':>6} {'Q*':>8}"
    )
    print(fmt)
    print("  " + "─" * (len(fmt) - 2))
    for _, r in res.iterrows():
        print(
            f"  {r['product']:<14} {r['price']:>6.0f} {r['u']:>6.1f} {r['o']:>6.1f} "
            f"{r['cf']:>7.4f} {r['z']:>7.4f} {r['demand_mean']:>6.0f} "
            f"{r['demand_std']:>6.0f} {r['q_star']:>8.1f}"
        )

    total_q = res["q_star"].sum()
    print(f"\n  Total production capacity required (no limit): {total_q:,.1f} units")

    return res


# ═════════════════════════════════════════════
# PART 2 — Capacity limit of 20,000 units
# ═════════════════════════════════════════════
def part_two(df: pd.DataFrame) -> tuple[float, pd.DataFrame]:
    sep(f"PART 2 — Optimal Production Quantity (Capacity = {CAPACITY_PART2:,} units)")

    theta, alloc = allocate(df, CAPACITY_PART2)

    print(f"\n  Shadow price θ = {theta:.6f}")
    print(f"\n  {'Product':<14} {'CF_constrained':>16} {'Q_constrained':>14}")
    print("  " + "─" * 48)
    for _, r in alloc.iterrows():
        print(f"  {r['product']:<14} {r['cf_constrained']:>16.6f} {r['q_constrained']:>14.1f}")

    total = alloc["q_constrained"].sum()
    print(f"\n  Total Q (should ≈ {CAPACITY_PART2:,}): {total:,.1f}")

    max_cf_row = alloc.loc[alloc["cf_constrained"].idxmax()]
    min_cf_row = alloc.loc[alloc["cf_constrained"].idxmin()]
    print(f"\n  Highest CF: {max_cf_row['product']}  ({max_cf_row['cf_constrained']:.6f})")
    print(f"    → Highest underage cost relative to total cost → produce most aggressively")
    print(f"  Lowest CF:  {min_cf_row['product']}  ({min_cf_row['cf_constrained']:.6f})")
    print(f"    → Lowest underage cost relative to total cost → produce most conservatively")

    return theta, alloc


# ═════════════════════════════════════════════
# PART 3 — Risk ranking + speculative capacity
# ═════════════════════════════════════════════
def part_three(df: pd.DataFrame, p2_theta: float, p2_alloc: pd.DataFrame) -> tuple:
    sep("PART 3 — Risk Ranking & Speculative Capacity Allocation")

    # Compute risk index for each product
    df_risk = df.copy()
    df_risk["u"] = df_risk["price"].apply(lambda p: UNDERAGE_PCT * p)
    df_risk["o"] = df_risk["price"].apply(lambda p: OVERAGE_PCT  * p)
    df_risk["risk_index"] = df_risk.apply(
        lambda r: risk_index(r["u"], r["o"], r["demand_mean"], r["demand_std"]), axis=1
    )
    df_risk["cv"] = df_risk["demand_std"] / df_risk["demand_mean"]
    df_risk = df_risk.sort_values("risk_index").reset_index(drop=True)

    print("\n  Risk Ranking (lowest = safest for speculative production):\n")
    print(f"  {'Rank':<6} {'Product':<14} {'Risk Index':>12} {'CV (σ/μ)':>10}")
    print("  " + "─" * 46)
    for rank, (_, r) in enumerate(df_risk.iterrows(), start=1):
        print(f"  {rank:<6} {r['product']:<14} {r['risk_index']:>12.4f} {r['cv']:>10.4f}")

    # ── Part 3A: Speculative capacity = 10,000 units ──────────────────────────
    print(f"\n  ── 3A: Speculative Capacity = {CAPACITY_PART3A:,} units ──")
    alloc_3a = _run_speculative(df_risk, CAPACITY_PART3A, p2_alloc, label="3A")

    # ── Part 3B: Speculative capacity = 5,000 units ───────────────────────────
    print(f"\n  ── 3B: Speculative Capacity = {CAPACITY_PART3B:,} units ──")
    alloc_3b = _run_speculative(df_risk, CAPACITY_PART3B, p2_alloc, label="3B")

    return df_risk, alloc_3a, alloc_3b


def _run_speculative(df_risk: pd.DataFrame, capacity: int,
                     p2_alloc: pd.DataFrame, label: str) -> pd.DataFrame:
    """
    Select low-risk products for speculative production.

    Walk the risk-ranked list and keep adding products as long as the
    *next* product's standalone Q* still fits.  If adding the next
    product would exceed capacity, include it anyway and let the
    shadow-price solver trim all selected products down to exactly
    meet the capacity constraint.  This is more realistic than leaving
    capacity on the table.
    """
    selected = []
    cumulative_q = 0.0
    for _, row in df_risk.iterrows():
        u, o = row["u"], row["o"]
        q    = q_star_normal(row["demand_mean"], row["demand_std"], u, o)
        if cumulative_q + q <= capacity:
            # Fits without binding the constraint
            selected.append(row)
            cumulative_q += q
        else:
            # Would exceed capacity — include it and let θ solve the trim
            selected.append(row)
            cumulative_q += q
            break  # stop here; remaining products are higher risk

    subset = pd.DataFrame(selected)

    theta, alloc = allocate(subset, capacity)

    # Build P2 CF lookup
    p2_cf = p2_alloc.set_index("product")["cf_constrained"].to_dict()

    print(f"\n  Selected products for speculative production ({len(subset)}):\n")
    print(f"  {'Product':<14} {'Risk Index':>12} {'CF (Part2)':>12} {'CF (Spec)':>11} {'Q_spec':>10}")
    print("  " + "─" * 63)
    for _, r in alloc.iterrows():
        cf2 = p2_cf.get(r["product"], float("nan"))
        print(
            f"  {r['product']:<14} "
            f"{df_risk.loc[df_risk['product']==r['product'], 'risk_index'].values[0]:>12.4f} "
            f"{cf2:>12.6f} "
            f"{r['cf_constrained']:>11.6f} "
            f"{r['q_constrained']:>10.1f}"
        )

    total = alloc["q_constrained"].sum()
    print(f"\n  Shadow price θ  = {theta:.6f}")
    print(f"  Total Q (should ≈ {capacity:,}): {total:,.1f}")

    excluded = [r["product"] for _, r in df_risk.iterrows()
                if r["product"] not in alloc["product"].values]
    print(f"\n  Products reserved for REACTIVE capacity: {', '.join(excluded)}")
    return alloc


# ═════════════════════════════════════════════
# PART 4 — Recommendations
# ═════════════════════════════════════════════
def part_four() -> None:
    sep("PART 4 — Capacity Management Recommendations")
    print("""
  1. PRODUCE LOW-RISK PRODUCTS SPECULATIVELY
     Products with low risk index (low CV, low cost consequences) can be produced
     early using forecast-based speculative capacity without significant downside.

  2. DELAY HIGH-RISK PRODUCTS UNTIL DEMAND IS KNOWN (REACTIVE CAPACITY)
     Products with high demand uncertainty (high CV) should wait for demand signals.
     Use overtime, subcontracting, or expedited sourcing as reactive capacity.

  3. INVEST IN EARLIER DEMAND SIGNALS
     Advance orders, pre-sales campaigns, and retailer commitments reduce forecast
     error and effectively increase reactive capacity without physical expansion.

  4. REDUCE LEAD TIMES FROM OVERSEAS FACTORY
     Shorter lead times compress the speculative window and let more production
     shift to the reactive phase — directly reducing overstock and stockout risk.

  5. DUAL SOURCING STRATEGY
     Pair the low-cost overseas factory (high volume, speculative) with a faster
     domestic or near-shore supplier (reactive top-up orders at higher unit cost
     but lower total risk). Shadow price θ quantifies the value of that extra capacity.

  6. SENSITIVITY ANALYSIS ON σ
     Small reductions in demand standard deviation (e.g. 10% improvement in
     forecasting accuracy) yield disproportionate reductions in required safety
     stock and optimal Q* buffers — especially for high-CV products.
    """)


# ═════════════════════════════════════════════
# EXPORT — Write results back into xlsx template
# ═════════════════════════════════════════════
#
# Template layout (openpyxl 1-indexed):
#
#   Row 2 :  B2:C2 (merged) = "Shadow Price for Capacity"  →  D2 = θ value
#   Row 4 :  Header row (B–K)
#   Rows 5–15 : 11 product rows
#   Row 16:  Total row
#
#   Col B (2)  Product Lines     — pre-filled
#   Col C (3)  Price             — pre-filled
#   Col D (4)  Demand Mean       — pre-filled
#   Col E (5)  Demand Std        — pre-filled
#   Col F (6)  Overage           — fill (accounting $)
#   Col G (7)  Underage          — fill (accounting $)
#   Col H (8)  Risk Ranking      — fill (0.0)
#   Col I (9)  Yes/No            — fill (text)
#   Col J (10) CF                — fill (0%)
#   Col K (11) Production Qty    — fill (integer 0)
#
# Strategy: write values only, preserving the template's existing cell
# formatting (fills, fonts, number formats, alignment).  The template
# already has the correct number formats on the empty answer cells.

_FIRST_DATA_ROW = 5
_TOTAL_ROW      = 16
_SHADOW_VALUE   = (2, 4)   # (row, col) — D2, right of the B2:C2 merged label

# Column indices (openpyxl 1-based)
_C_OVERAGE   = 6   # F
_C_UNDERAGE  = 7   # G
_C_RISK      = 8   # H
_C_YESNO     = 9   # I
_C_CF        = 10  # J
_C_Q         = 11  # K


def export_to_xlsx(
    p1: pd.DataFrame,
    p2_theta: float,
    p2_alloc: pd.DataFrame,
    df_risk: pd.DataFrame,
    speculative_10k: list[str],
) -> Path:
    """
    Copy the original FLA template and fill every blank answer column
    with computed results, preserving the template's formatting.

    Returns the path to the saved file.
    """
    src  = Path(__file__).parent / "data" / "bana6420_u3_assignment-finger-lakes-data.xlsx"
    dest = Path(__file__).parent / "data" / "BANA6420_Gill,Nathanael_Unit3Assignment.xlsx"
    shutil.copy2(src, dest)

    wb = openpyxl.load_workbook(dest)
    ws = wb["Data"]

    # ── Build lookup dicts keyed by product name ──────────────────────────
    p1_lookup   = p1.set_index("product").to_dict("index")
    p2_lookup   = p2_alloc.set_index("product").to_dict("index")
    risk_lookup = df_risk.set_index("product")["risk_index"].to_dict()
    spec_set    = set(speculative_10k)

    # Product order matches the sheet (row 5 → row 15)
    products = p1["product"].tolist()

    for i, product in enumerate(products):
        row = _FIRST_DATA_ROW + i
        p1r = p1_lookup[product]
        p2r = p2_lookup[product]

        # F — Overage cost
        ws.cell(row, _C_OVERAGE).value  = p1r["o"]
        # G — Underage cost
        ws.cell(row, _C_UNDERAGE).value = p1r["u"]
        # H — Risk ranking index
        ws.cell(row, _C_RISK).value     = risk_lookup[product]
        # I — Yes / No (speculative @ 10k)
        ws.cell(row, _C_YESNO).value    = "Yes" if product in spec_set else "No"
        # J — Critical Fractile (constrained, Part 2)
        ws.cell(row, _C_CF).value       = p2r["cf_constrained"]
        # K — Production Quantity (constrained, Part 2)
        ws.cell(row, _C_Q).value        = round(p2r["q_constrained"])

    # ── Total row (K16) — SUM formula so it stays live ────────────────────
    ws.cell(_TOTAL_ROW, _C_Q).value = f"=SUM(K{_FIRST_DATA_ROW}:K{_FIRST_DATA_ROW + len(products) - 1})"

    # ── Shadow price θ (D2) ───────────────────────────────────────────────
    ws.cell(*_SHADOW_VALUE).value = round(p2_theta, 6)

    wb.save(dest)
    return dest


# ═════════════════════════════════════════════
# Main
# ═════════════════════════════════════════════
if __name__ == "__main__":
    # ── Load data ────────────────────────────────────────────────────────────
    df = load_finger_lakes()
    print("Finger Lakes data loaded successfully.")
    print(df.to_string(index=False))

    # Attach costs to df for reuse across parts
    df["u"] = df["price"].apply(lambda p: UNDERAGE_PCT * p)
    df["o"] = df["price"].apply(lambda p: OVERAGE_PCT  * p)

    # ── Run all parts ─────────────────────────────────────────────────────────
    p1_results          = part_one(df)
    p2_theta, p2_alloc  = part_two(df)
    df_risk, alloc_3a, alloc_3b = part_three(df, p2_theta, p2_alloc)
    part_four()

    # ── Export to xlsx ────────────────────────────────────────────────────────
    speculative_10k = alloc_3a["product"].tolist()
    out = export_to_xlsx(
        p1              = p1_results,
        p2_theta        = p2_theta,
        p2_alloc        = p2_alloc,
        df_risk         = df_risk,
        speculative_10k = speculative_10k,
    )
    print(f"\n✅  Results exported to: {out}")
