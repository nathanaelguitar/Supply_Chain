# populate_excel_template.py - U2: Populate Excel template with calculated results
import pandas as pd
import openpyxl
from pathlib import Path
from u2_inventory_optimization import *

def populate_excel_template():
    """Populate the Cosmos Excel template with calculated values."""
    
    # Load the template
    data_dir = Path(__file__).parent / "data"
    template_path = data_dir / 'bana6420_u2_assignment-cosmos-data.xlsx'
    output_path = Path(__file__).parent / 'BANA6420_Gill,Nathanael_Unit2Assignment.xlsx'
    
    # Load workbook
    wb = openpyxl.load_workbook(template_path)
    ws = wb['Data']
    
    # Run calculations first
    results = part_one_calculations()
    
    # Service Target (Row 3)
    ws['C3'] = SERVICE_LEVEL_TARGET
    
    # Column mapping: RDC=C, DC1(Albany)=D, DC2(Brooklyn)=E, DC3(Syracuse)=F
    locations = {
        'C': 'Brooklyn RDC',
        'D': 'Albany DC', 
        'E': 'Brooklyn DC',
        'F': 'Syracuse DC'
    }
    
    # Row 7: Daily Demand Mean
    ws['D7'] = DEMAND_ALBANY['mean']
    ws['E7'] = DEMAND_BROOKLYN_DC['mean']
    ws['F7'] = DEMAND_SYRACUSE['mean']
    ws['C7'] = results['Brooklyn RDC']['daily_demand_mean']
    
    # Row 8: Daily Demand Std
    ws['D8'] = DEMAND_ALBANY['std']
    ws['E8'] = DEMAND_BROOKLYN_DC['std']
    ws['F8'] = DEMAND_SYRACUSE['std']
    ws['C8'] = results['Brooklyn RDC']['daily_demand_std']
    
    # Row 9: Lead Time Mean (days)
    ws['D9'] = LT_ALBANY_DC['mean']
    ws['E9'] = LT_BROOKLYN_DC['mean']
    ws['F9'] = LT_SYRACUSE_DC['mean']
    ws['C9'] = LT_RDC_FROM_PLANT['mean']
    
    # Row 10: Lead Time Std (days)
    ws['D10'] = LT_ALBANY_DC['std']
    ws['E10'] = LT_BROOKLYN_DC['std']
    ws['F10'] = LT_SYRACUSE_DC['std']
    ws['C10'] = LT_RDC_FROM_PLANT['std']
    
    # Row 11: Review Period (days)
    for col in ['C', 'D', 'E', 'F']:
        ws[f'{col}11'] = REORDER_PERIOD
    
    # Row 12: Exposure Period Demand Mean (lambda * (R+L))
    for col, loc in locations.items():
        data = results[loc]
        ws[f'{col}12'] = data['daily_demand_mean'] * (data['reorder_period'] + data['lead_time_mean'])
    
    # Row 13: Exposure Period Demand Std (sigma_R_plus_L)
    for col, loc in locations.items():
        ws[f'{col}13'] = results[loc]['sigma_R_plus_L']
    
    # Row 15: Cycle Stock
    for col, loc in locations.items():
        data = results[loc]
        cycle_stock = calculate_cycle_stock(data['daily_demand_mean'], data['reorder_period'])
        ws[f'{col}15'] = cycle_stock
    
    # Row 16: Pipeline Stock
    for col, loc in locations.items():
        data = results[loc]
        pipeline_stock = calculate_pipeline_stock(data['daily_demand_mean'], data['lead_time_mean'])
        ws[f'{col}16'] = pipeline_stock
    
    # Row 17: Safety Stock
    for col, loc in locations.items():
        ws[f'{col}17'] = results[loc]['safety_stock']
    
    # Row 18: Total Inventory
    for col in ['C', 'D', 'E', 'F']:
        ws[f'{col}18'] = f'=SUM({col}15:{col}17)'
    
    # Row 19: Unit Value
    ws['C19'] = RDC_UNIT_COST
    for col in ['D', 'E', 'F']:
        ws[f'{col}19'] = DC_UNIT_COST
    
    # Row 20: Annual Holding Cost Rate
    for col in ['C', 'D', 'E', 'F']:
        ws[f'{col}20'] = ANNUAL_HOLDING_RATE
    
    # Row 21: Annual Inventory Holding Cost
    for col in ['C', 'D', 'E', 'F']:
        ws[f'{col}21'] = f'={col}18*{col}19*{col}20'
    
    # Row 23: Unit Transportation Cost
    ws['C23'] = TP_PLANT_TO_RDC
    ws['D23'] = TP_RDC_TO_ALBANY
    ws['E23'] = TP_RDC_TO_BROOKLYN
    ws['F23'] = TP_RDC_TO_SYRACUSE
    
    # Row 24: Annual Demand
    for col, loc in locations.items():
        data = results[loc]
        ws[f'{col}24'] = data['daily_demand_mean'] * 365
    
    # Row 25: Annual Transportation Cost
    for col in ['C', 'D', 'E', 'F']:
        ws[f'{col}25'] = f'={col}23*{col}24'
    
    # Row 27: Annual Inv + Transport Cost
    for col in ['C', 'D', 'E', 'F']:
        ws[f'{col}27'] = f'={col}21+{col}25'
    
    # Row 27 Column G (System Total)
    ws['G27'] = '=SUM(C27:F27)'
    
    # Save the populated workbook
    wb.save(output_path)
    print(f"\n✅ Excel template populated and saved to:")
    print(f"   {output_path}")
    print(f"\nFile naming convention: BANA6420_Gill,Nathanael_Unit2Assignment.xlsx")
    
    return output_path

if __name__ == "__main__":
    populate_excel_template()
